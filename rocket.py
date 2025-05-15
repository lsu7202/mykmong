from PySide6.QtWidgets import *
from PySide6.QtCore import *
from PySide6.QtGui import QAction, QPixmap, QIcon
import os, math, ast, requests, json, re, sys
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import msoffcrypto
import pandas as pd
from io import BytesIO
from getmac import *
from PIL import Image, ImageDraw
from PIL import ImageFont
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from googletrans import Translator
from itertools import permutations

#EEF1F3 #EBEDEF

class Caches():
    
    def makeCachesFolder(self):
        newFolder = "Caches"
        os.makedirs(newFolder, exist_ok=True)

        newFile = os.path.join(newFolder, "topImgPath.txt")
        if not os.path.exists(newFile):
            with open(newFile, "w", encoding="utf-8") as file:
                file.write("상단이미지")

        newFile = os.path.join(newFolder, "saveFolderPath.txt")
        if not os.path.exists(newFile):
            with open(newFile, "w", encoding="utf-8") as file:
                file.write("/")

        newFile = os.path.join(newFolder, "keywords.txt")
        if not os.path.exists(newFile):
            with open(newFile, "w", encoding="utf-8") as file:
                file.write("")

    def makeThumbFolder(Self, title):
        newFolder = f"썸네일/{title}"
        os.makedirs(newFolder, exist_ok=True)
        return newFolder
    def makeDeImgFolder(self, title):
        newFolder = f"상세이미지/{title}"
        os.makedirs(newFolder, exist_ok=True)
        return newFolder

        
    def getKeywordsCache(self):
        with open("Caches/keywords.txt", "r", encoding="utf-8") as file:
            content = file.read()
            data = ast.literal_eval(content)
        return data
    
    def getSupplyCache(self):
        with open("Caches/supply.txt", "r", encoding="utf-8") as file:
            content = file.read()
            data = ast.literal_eval(content)[0]
        return data
    
    def getExcRatingCache(self):
        with open("Caches/supply.txt", "r", encoding="utf-8") as file:
            content = file.read()
            data = ast.literal_eval(content)[1]
        return data

    def getCoupPriceCache(self): #판매가배수
        with open("Caches/supply.txt", "r", encoding="utf-8") as file:
            content = file.read()
            data = ast.literal_eval(content)[2]
        return data
        

class LoadExcelData():
     
    def loadData(self, filePath):
        """ 엑셀 파일을 불러와 '상품명 → 옵션1' 기준으로 2단계 그룹화 """
        df = pd.read_excel(filePath)
        grouped = df.groupby(["상품명", "옵션1"])  # 2단계 그룹화
        return grouped
    
    def getThumbName(self, groupedData, product_name, option1):
        """ 상품명에 맞는 썸네일 데이터 반환 """
        return list(set(optionImg for (product, opt1), df in groupedData if product == product_name and opt1 == option1 for optionImg in df["썸네일"]))
        
 
    def getTitleList(self, groupedData):
        """ 엑셀에서 상품명 리스트를 원래 순서대로 반환 (NaN 제거 후 중복 제거) """
        return list(dict.fromkeys(
            product for product, _ in groupedData.groups.keys() if pd.notna(product)
        ))

    def getOptionImageList(self, groupedData, product_name, option1):
        """ 특정 상품명과 옵션1 값에 해당하는 옵션이미지 리스트 반환 (중복 제거) """
        return list(set(optionImg for (product, opt1), df in groupedData if product == product_name and opt1 == option1 for optionImg in df["옵션이미지"]))
    
    def getOption1List(self, groupedData, product_name):
        """ 특정 상품의 옵션1 리스트 반환 (중복 제거) """
        return list(set(option for (product, option) in groupedData.groups.keys() if product == product_name))

    def getOption2List(self, groupedData, product_name, option1):
        """ 특정 상품과 옵션1에 대한 옵션2 리스트 반환 (중복 제거) """
        return list(set(option2 for (product, opt1), df in groupedData if product == product_name and opt1 == option1 for option2 in df["옵션2"]))
    
    def loadDataV2(self, groupedData):
        """ 기존 '상품명 → 옵션1' 그룹화된 데이터를 받아 '옵션2 → 옵션3'까지 추가 그룹화 """
        
        new_grouped = []
        
        for (product, opt1), df in groupedData:
            # 옵션2, 옵션3이 없는 경우 빈 문자열로 처리
            df[["옵션2", "옵션3"]] = df[["옵션2", "옵션3"]].fillna("")
            
            # 옵션2와 옵션3까지 포함해서 새로운 df 만들기
            for (opt2, opt3), sub_df in df.groupby(["옵션2", "옵션3"]):
                # 각 그룹을 새로운 형태로 추가 (상품명, 옵션1, 옵션2, 옵션3 형태로)
                if opt2 == "" and opt3 == "":
                    # 옵션2, 옵션3이 모두 없으면 (상품명, 옵션1) 형태로 추가
                    new_grouped.append((product, opt1))
                elif opt3 == "":
                    # 옵션3이 없으면 (상품명, 옵션1, 옵션2) 형태로 추가
                    new_grouped.append((product, opt1, opt2))
                else:
                    # 옵션3까지 있으면 (상품명, 옵션1, 옵션2, 옵션3) 형태로 추가
                    new_grouped.append((product, opt1, opt2, opt3))
        
        return new_grouped

    
    def getCombinedOptionList_fixed(self, groupedData, product_name):
        """ 특정 상품에 대한 옵션1 + 옵션2 + 옵션3 리스트를 반환 (중복 제거) """
        combined_options = []

        for df in list(groupedData):
            if df[0] == product_name:
            
                if len(df) == 2: # 옵션1만
                    combined_options.append(df[1])

                elif len(df) == 3: #옵션2까지
                    combined_options.append(f"{df[1]}분리선{df[2]}")
                elif len(df) == 4: #옵션3까지
                    combined_options.append(f"{df[1]}분리선{df[2]}분리선{df[3]}")

        return list(dict.fromkeys(combined_options))
            

        for (product, opt1, opt2, opt3), df in groupedData:
            if product == product_name:
                # 옵션2, 옵션3이 비어있을 경우 대비해 안전하게 처리
                opt2 = opt2 if opt2 else ""
                opt3 = opt3 if opt3 else ""

                # 옵션 조합 만들기 (불필요한 분리선 제거)
                option_str = opt1
                if opt2:
                    option_str += f"분리선{opt2}"
                if opt3:
                    option_str += f"분리선{opt3}"

                combined_options.append(option_str)

        return list(dict.fromkeys(combined_options))  # 중복 제거


    def getCombinedOptionList(self, groupedData, product_name):
        """ 특정 상품에 대한 옵션1 + 옵션2 + 옵션3 리스트를 엑셀 순서대로 반환 (중복 제거) """

        
        combined_options = []

        # 1차 그룹: 옵션1 + 옵션2 조합 만들기
        temp_options = {}  # {(옵션1, 옵션2): []} 형태로 저장
        for (product, opt1), df in groupedData:
            if product == product_name:
                for _, row in df.iterrows():
                    option2 = row["옵션2"] if pd.notna(row["옵션2"]) else ""
                    option3 = row["옵션3"] if "옵션3" in df.columns and pd.notna(row["옵션3"]) else ""

                    key = (opt1, option2)
                    if key not in temp_options:
                        temp_options[key] = []

                    # 옵션3이 있으면 추가
                    if option3:
                        temp_options[key].append(option3)
                    elif option2 == "":  # 옵션2가 비어있고 옵션3이 없을 경우
                        # 옵션1만 있는 경우도 처리
                        combined_options.append(f"{opt1}분리선")
            
        

        # 2차 그룹: 옵션1 + 옵션2 + 옵션3 조합 만들기
        for (opt1, opt2), option3_list in temp_options.items():
            if option3_list:  # 옵션3이 있을 경우
                for opt3 in option3_list:
                    combined_options.append(f"{opt1}분리선{opt2}분리선{opt3}분리선")
            else:  # 옵션3이 없는 경우
                combined_options.append(f"{opt1}분리선{opt2}")

        return list(dict.fromkeys(combined_options))  # 중복 제거

    

    def getCost(self, groupedData, product_name, option):
        """상품명과 옵션1, 옵션2, 옵션3에 맞는 '원가' 속성 반환"""

        options = option.split("분리선")
        options = [item for item in options if item not in ("", " ")]

        # 옵션1, 옵션2, 옵션3을 나누기
        if len(options) == 3:
            option1 = options[0].strip()
            option2 = options[1].strip()
            option3 = options[2].strip()

        elif len(options) == 2:
            option1 = options[0].strip()
            option2 = options[1].strip()
            option3 = None

        elif len(options) == 1:
            option1 = options[0].strip()
            option2 = None
            option3 = None

        cost = []

        # 옵션1, 옵션2, 옵션3이 맞는 행을 찾아서 원가를 추출
        for (product, opt1), df in groupedData:
            if product.strip() == product_name.strip() and opt1.strip() == option1.strip():
                for _, row in df.iterrows():
                    # 옵션2가 None이거나 일치하는 경우
                    if (option2 is None or row["옵션2"].strip() == option2) and \
                    (option3 is None or row["옵션3"].strip() == option3):
                        cost.append(row["원가"])

        # 원가 값을 반환 (없으면 None 반환)
        return float(cost[0]) if cost else None  # 첫 번째 원가 반환



class run1688(QWidget):
    def __init__(self):
        super().__init__()
        self.scrapGUI()
    
    def scrapGUI(self):
        self.driver = self.getDrivers()
        self.myDriver()
        self.setWindowTitle("")
        self.setGeometry(100,100,405,380)
        self.setStyleSheet("background-color: #212121; color: white")

        # 메인 레이아웃 설정
        main_layout = QVBoxLayout()
        
        saveLogInfoBtn = QPushButton("로그인 정보 저장")
        saveLogInfoBtn.setStyleSheet("""
            QPushButton {
                font-size: 20px;
                background-color: #A0A0A5;
                color: white;
                border-radius: 120px;  /* 반지름을 버튼 크기의 절반으로 설정 */
            }
            QPushButton:hover {
                background-color: #FFA500;
            }
        """)

        saveLogInfoBtn.clicked.connect(self.loginTaoBao)

        main_layout.addWidget(saveLogInfoBtn)

        self.myUrl = QLineEdit()
        self.myUrl.setStyleSheet("background-color: black; color: white")
        self.myUrl.setPlaceholderText("URL 입력")
        main_layout.addWidget(self.myUrl)


        # Run 버튼
        self.run_button = QPushButton("")
        self.run_button.setFixedSize(280, 280)  # 정사각형 크기로 설정

        self.run_button.setStyleSheet("""
            QPushButton {
                font-size: 20px;
                background-color: #A0A0A5;
                color: white;
                border-radius: 120px;  /* 반지름을 버튼 크기의 절반으로 설정 */
            }
            QPushButton:hover {
                background-color: #FFA500;
            }
        """)
        self.run_button.setCursor(Qt.PointingHandCursor)
        self.run_button.clicked.connect(self.getDataSheet)

        self.setMaxThumb = QLineEdit()
        self.setMaxThumb.setStyleSheet("background-color: #EBEDEF; color: black;")
        self.setMaxThumb.setPlaceholderText("상세페이지 최대 이미지 수")


        scroll_area = QScrollArea()
        scroll_area.setMinimumHeight(160)
        scroll_widget = QWidget()
        self.ThumbLayout = QHBoxLayout(scroll_widget)

        scroll_widget.setLayout(self.ThumbLayout)
        scroll_area.setWidget(scroll_widget)
        scroll_area.setWidgetResizable(True)

        main_layout.addWidget(self.run_button, alignment=Qt.AlignCenter)
        main_layout.addWidget(self.setMaxThumb)
        main_layout.addWidget(scroll_area)

        self.concatBtn = QPushButton("조합")
        self.concatBtn.setStyleSheet("""
            QPushButton {
                background-color: #A0A0A5;
                color: white;
            }
            QPushButton:hover {
                background-color: #FFA500;
            }
        """)
        self.concatBtn.clicked.connect(self.concatImg)

        main_layout.addWidget(self.concatBtn)

        self.setLayout(main_layout)

    def loginTaoBao(self):
        self.driver.refresh()
        cookies = self.driver.get_cookies()
        with open("logincookies.json", "w") as f:
                json.dump(cookies, f)

    
    def setSavePath(self):
        folderPath = QFileDialog.getExistingDirectory(self, "폴더 선택", "")
        
        if folderPath:
            with open("Caches/saveFolderPath.txt", "w", encoding="utf-8") as file:
                file.write(folderPath)

    def trans(self, text):
        translator = Translator()
        translated = translator.translate(text, src='zh-cn', dest="ko")
        return translated.text
    
    def getDrivers(self):
        # Chrome 옵션 설정
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0")

        #chrome_options.add_argument("profile-directory=default")  # 기본 프로필 사용 (선택 사항)
        return webdriver.Chrome(options=chrome_options)
    
    def getHeaders(self):
        # 크롬 헤더, 확인 필요
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0",
            "Referer": "https://www.1688.com/",
            "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        }
        return headers

    # Chrome 개발자 도구에서 복사한 쿠키 추가 (로그인이 필요한 경우)
    def getCookies(self):
        file_path = 'cookies.json'
        # 존재 여부 확인
        if os.path.exists(file_path):
            print("쿠키파일 읽기")
            with open("cookies.json", "r") as f:
                cookies = json.load(f)
                return cookies
        else:
            print("쿠키 다운")
            self.driver.refresh()
            cookies = {cookie['name']: cookie['value'] for cookie in self.driver.get_cookies()}
            cookies = {'cna' : cookies["cna"], 'cookie2' : cookies["cookie2"], '_tb_token_' : cookies["_tb_token_"], 'isg': cookies["isg"]}
            with open("cookies.json", "w") as f:
                json.dump(cookies, f)
            return cookies

    def myDriver(self):
        file_path = 'logincookies.json'
        # 존재 여부 확인
        if os.path.exists(file_path):
            with open(file_path, "r") as f:
                cookies = json.load(f)
        
        # 1688 접속
        url = "https://1688.com"
        self.driver.get(url)
        
        if os.path.exists(file_path):
            for cookie in cookies:
                self.driver.add_cookie(cookie)
        
        self.driver.refresh()
        

    def getData(self):
        self.driver.implicitly_wait(10)
        # 가장 마지막 탭에 페이지 포커스
        
        
        # requests 할 url, headers, cookies
        url = self.myUrl.text()
        headers = self.getHeaders()
        cookies = self.getCookies()

        # 현재 페이지 주소로 HTML 요청
        response = requests.get(url, headers=headers, cookies=cookies)

        # 응답이 정상인지 확인
        if response.status_code == 200:
            html_source = response.text
        else:
            print("❌ 다운로드 실패:", response.status_code)

        soup = BeautifulSoup(html_source, "html.parser")

        # 7번째 <script> 태그 선택 (여기서 script_tags[7]로 지정)
        script_tags = soup.find_all("script")
        script_tag = script_tags[7].string

        with open("scripttag.txt", "w", encoding="utf-8") as file:
            file.write(script_tag)

        cleaned_text = re.sub(r'window\.__GLOBAL_DADA=\{.*?\};', '', script_tag)
        cleaned_text = re.sub(r'window\.__INIT_DATA=', '', cleaned_text)

        data_dict = json.loads(cleaned_text)

        return data_dict
      
    def getDataSheet(self):

        #데이터 시트
        self.dataSheet = []
        option1 = " "
        option2 = " "
        cost = 0
        thumb_text = ""
    
        # html 데이터 로딩
        myDict = self.getData()
        row = {"상품명" : "", "옵션1" : " ", "옵션2" : " ", "원가" : 0, "옵션이미지" : ""}

        topImgPath = "상단이미지/상단이미지.001.png"
        self.topImg = Image.open(topImgPath)

        def getTitle(data):
            title = data['data']['1081181309095']['data']['title']
            return self.trans(title)
        
        title = getTitle(myDict)

        def getPriceType(data):
            return data['globalData']['orderParamModel']['orderParam']['skuParam']['skuPriceType']
        
        def downDetailImg(detailImgUrl, title, option, idx):
            folderPath = Caches().makeThumbFolder(title)
            try:
                response = requests.get(detailImgUrl, stream=True, headers=self.getHeaders())
                response.raise_for_status()  # 오류 확인
                
                # 다운로드한 이미지 데이터를 PIL로 열기
                with Image.open(response.raw) as image:
                    # 이미지 형식 추출
                    file_ext = image.format.lower()  # 'JPEG', 'PNG', 'GIF' 등으로 추출됨
                    print(f"Image format detected: {file_ext}")
                    
                    # 파일명 생성
                    thumb_text = f"{title}_{option}_{idx}".replace("/", "").replace(" ", "")
                    file_path = os.path.join(folderPath, f"{thumb_text}.{file_ext}")
                    
                    
                    # 이미지 저장
                    with open(file_path, "wb") as file:
                        # 이미지 데이터를 파일로 저장
                        image.save(file, format=image.format)
                        
            except requests.exceptions.RequestException as e:
                print(f"다운로드 실패: {imgUrl}, 오류: {e}")

            return file_path
        
        def downThumbImg(thumbImgUrl, title, idx):
            folderPath = Caches().makeDeImgFolder(title)
            try:
                # 파일 확장자 추출
                file_ext = imgUrl.split('.')[-1].split('?')[0]  # URL에 ?가 있을 경우 대비
                if file_ext == "gif":
                    return False
                # 상품명_옵션명.jpg, '/' 랑 '.' 있으면 삭제 
                thumb_text = f"{title}_썸네일_{idx}".replace("/", "").replace(".", "")
                file_path = os.path.join(folderPath, f"{thumb_text}.{file_ext}")

                # 이미지 다운로드
                response = requests.get(thumbImgUrl, stream=True, headers=self.getHeaders())
                response.raise_for_status()  # 오류 확인

                # 이미지 저장
                with open(file_path, "wb") as file:
                    for chunk in response.iter_content(1024):
                        file.write(chunk)

                image = Image.open(file_path)
                #resized = image.resize((500,500))
                image.save(file_path)

                
            except requests.exceptions.RequestException as e:
                print(f"다운로드 실패: {imgUrl}, 오류: {e}")

            return file_path
        

        # skuInfoMap 으로 경로 이동
        items = myDict['globalData']['skuModel']['skuInfoMap']
        # skuProps 으로 경로 이동
        detailLinks = myDict['globalData']['skuModel']['skuProps'][0]['value']
        # images 으로 경로 이동
        thumbLinks = myDict['globalData']['images'] # 리스트형태

        self.detailImglist = [] #concatImg에서 사용
        for idx, (img) in enumerate(detailLinks):
            row = {"상품명" : "", "옵션1" : " ", "옵션2" : " ", "원가" : 0, "옵션이미지" : ""}
            if 'name' in img:
                imgoption = self.trans(img['name'])
                
            if 'imageUrl' in img:
                imgUrl = (img['imageUrl'])
                #옵션 이미지 다운
                self.detailImglist.append(downDetailImg(imgUrl, title , imgoption, idx))

        thumbLinksList = []
        i=0 #썸네일 번호[인덱스]
        for img in thumbLinks:
            
            if 'searchImageURI' in img:
                imgUrl = (img['searchImageURI'])
                ThumbImg = downThumbImg(imgUrl, title, i)
                if ThumbImg:
                    thumbLinksList.append(ThumbImg)
                i += 1
        
        self.checkboxList = {}
        for thumb in thumbLinksList:
            checkbox = QCheckBox()
            pixmap = QPixmap(thumb)
            icon = QIcon(pixmap)
            checkbox.setIcon(icon)
            checkbox.setIconSize(QSize(120,120))
            self.checkboxList[checkbox] = thumb
            self.ThumbLayout.addWidget(checkbox)

        # 원하는 키, 값 검색
        for info in items.values():
            row = {"상품명" : title, "옵션1" : "", "옵션2" : "", "원가" : 0, "옵션이미지" : ""}

            option1 = self.trans(info['specAttrs'].strip().replace("&gt", ""))

            if ";" in option1:
                parts = (option1.split(";"))

                option2 = parts[1]
                option1 = parts[0]

            for detail in self.detailImglist:
                p = detail.split("_")
                if p[1] == option1.replace(" ", ""):

                    idx = p[2]

            thumb_text = f"{title}_{option1}_{idx}".replace("/","").replace(" ","")

            row["옵션1"] = option1
            if option2:
                row["옵션2"] = option2
            else:
                row["옵션2"] = " "

            row["옵션이미지"] = thumb_text

            if getPriceType(myDict) == "skuPrice":
                price = info['price']
                #print(f"원가 : {price}")
                cost = float(price)
                row["원가"] = cost

            elif getPriceType(myDict) == "rangePrice":
                price = myDict['globalData']['orderParamModel']['orderParam']['skuParam']["skuRangePrices"][0]['price']
                #원가
                cost = float(price)
                row["원가"] = cost

            self.dataSheet.append(row)

        self.saveExl()

        self.title_self = title
    
    def clear_layout(self, layout):
        # 레이아웃의 모든 위젯 제거
        if layout is not None:
            while layout.count():
                item = layout.takeAt(0)  # 첫 번째 아이템을 가져옴
                if item.widget():
                    item.widget().deleteLater()

    def concatImg(self):
        topImg = Image.open("상단이미지/상단이미지.001.png")
        title = self.title_self
        folderPath = Caches().makeDeImgFolder(title)
        
        # 옵션 이미지 불러오기
        deimages = [Image.open(img) for img in self.detailImglist]
        if self.setMaxThumb.text():
            max_idx = int(self.setMaxThumb.text())
            deimages = deimages[:max_idx]
        
        img_width, img_height = deimages[0].size
        
        # 상단 이미지 사이즈 조절 (가로 크기를 기준으로 조정)
        new_width = img_width * 2 + 15
        width_percent = new_width / float(topImg.size[0])
        new_height = int(float(topImg.size[1]) * width_percent)
        topImg = topImg.resize((new_width, new_height), Image.LANCZOS)
        
        # 체크된 썸네일 이미지 가져오기
        thumblist = [self.checkboxList[cb] for cb in self.checkboxList if cb.isChecked()]
        thimages = [Image.open(img) for img in thumblist]
        
        # 썸네일 이미지 크기 조정 (topImg의 너비에 맞춤)
        thumb_width = topImg.width
        for i in range(len(thimages)):
            scale = thumb_width / thimages[i].width
            new_thumb_height = int(thimages[i].height * scale)
            thimages[i] = thimages[i].resize((thumb_width, new_thumb_height), Image.LANCZOS)
        
        # 썸네일 높이 계산
        total_thumb_height = sum(img.height for img in thimages)
        
        # 옵션 이미지 배열 설정 (2열 배치)
        cols = 2
        rows = math.ceil(len(deimages) / cols)
        gap = 15
        text_height = 15  # 텍스트 높이
        
        de_width = (img_width + gap) * cols - gap
        de_height = (img_height + text_height + gap) * rows - gap
        
        # 최종 캔버스 크기 계산
        final_width = max(topImg.width, de_width)
        final_height = topImg.height + total_thumb_height + de_height
        final_img = Image.new("RGB", (final_width, final_height), "white")
        
        # 상단 이미지 배치
        y_offset = 0
        final_img.paste(topImg, (0, y_offset))
        y_offset += topImg.height
        
        # 썸네일 배치
        for img in thimages:
            final_img.paste(img, (0, y_offset))
            y_offset += img.height
        
        # 옵션 이미지 배치 (2열)
        draw = ImageDraw.Draw(final_img)
        try:
            font = ImageFont.load_default()
        except:
            print("no font")
        
        for idx, img in enumerate(deimages):
            x_offset = (idx % cols) * (img_width + gap)
            y_position = y_offset + (idx // cols) * (img_height + text_height + gap)
            
            final_img.paste(img, (x_offset, y_position))
            
            # 파일명 텍스트 추가 (예제용으로 빈 문자열)
            filename = ""
            text_x = x_offset + 10
            text_y = y_position + img_height + 5
            draw.text((text_x, text_y), filename, font=font, fill="black")
        
        # 이미지 크기 조정 후 저장
        final_img = self.resize_image(final_img)
        final_img.save(f"{folderPath}/{title}_상세.jpg")
        self.clear_layout(self.ThumbLayout)

    def resize_image(self, img, min_width=780, min_height=500):
        original_width, original_height = img.size
        if original_width >= min_width and original_height >= min_height:
            return img
        
        scale_w = min_width / original_width
        scale_h = min_height / original_height
        scale = max(scale_w, scale_h)
        new_width = int(original_width * scale)
        new_height = int(original_height * scale)
        
        return img.resize((new_width, new_height), Image.LANCZOS)


    def saveExl(self):
        df = pd.DataFrame(self.dataSheet)
        # 엑셀 파일로 저장 (ExcelWriter 사용하면 여러 시트 저장 가능)
        if os.path.exists("상품_데이터.xlsx"):
            with pd.ExcelWriter("상품_데이터.xlsx", mode="a", engine='openpyxl', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, index=False, header=False, startrow=writer.sheets["Sheet1"].max_row)
        else:
            # 파일이 없으면 새로 저장
            df.to_excel("상품_데이터.xlsx", index=False, engine="openpyxl")
            
class DropLineEdit(QLineEdit):
    pathDropped = Signal(str)
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls:
                path = urls[0].toLocalFile()
                self.setText(path)
                self.pathDropped.emit(path)

        else:
            event.ignore()

class MainWindow(QMainWindow):

    def __init__(self):
        super().__init__()

        path = "Caches/saveFolderPath.txt"
    
        # 파일이 없거나 비어 있으면 데스크탑 경로 저장
        if not os.path.exists(path) or os.path.getsize(path) == 0:
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            with open(path, "w") as file:
                file.write(desktop_path)
                self.loadedPath = desktop_path

        else:
            with open(path, "r") as file:
                self.loadedPath = file.read()
            
        self.mainGUI()
        

    def mainGUI(self):
        self.setWindowTitle("")
        self.setGeometry(100,100,620,700)
        self.setStyleSheet("background-color: #212121; color: white")

        mainWidget = QWidget()
        mainWidget.setStyleSheet("background-color: #212121; color: #d8d8d8")
        self.setCentralWidget(mainWidget)

        mainLayout = QVBoxLayout()
        mainWidget.setLayout(mainLayout)                      


        run1688Btn = QPushButton("Run-1688")
        run1688Btn.setStyleSheet("""
            QPushButton {
                background-color: #303030;
                color: #d8d8d8;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FFA500;
            }
        """)
        run1688Btn.clicked.connect(self.openRun1688)

        loadSavePathLayout = QHBoxLayout()
        self.loadSavePath = QLineEdit()
        self.loadSavePath.setStyleSheet("background-color: #303030; color: #d8d8d8;")
        self.loadSavePath.setReadOnly(True)
        self.loadSavePath.setText(self.loadedPath)

        loadPathBtn = QPushButton("찾아보기")
        loadPathBtn.setStyleSheet("""
            QPushButton {
                background-color: #303030;
                color: #d8d8d8;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FFA500;
            }
        """)
        loadPathBtn.clicked.connect(self.loadsvPath)
        loadSavePathLayout.addWidget(QLabel("저장경로"))
        loadSavePathLayout.addWidget(self.loadSavePath)
        loadSavePathLayout.addWidget(loadPathBtn)

        makeNewFolderLayout = QHBoxLayout()
        self.folderName = QLineEdit()
        self.folderName.setText("새 폴더")
        self.folderName.setStyleSheet("background-color: #303030; color: #d8d8d8;")
        makeBtn = QPushButton("생성")
        makeBtn.setStyleSheet("""
            QPushButton {
                background-color: #303030;
                color: #d8d8d8;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FFA500;
            }
        """)
        makeBtn.clicked.connect(self.makeFolder)

        makeNewFolderLayout.addWidget(QLabel("폴더생성하기"))
        makeNewFolderLayout.addWidget(self.folderName)
        makeNewFolderLayout.addWidget(makeBtn)

        loadExlLayout = QHBoxLayout()

        self.loadExlBrowser = DropLineEdit()
        self.loadExlBrowser.setStyleSheet("background-color: #303030; color: #d8d8d8;")
        self.loadExlBrowser.setReadOnly(True)
        self.loadExlBrowser.pathDropped.connect(self.loadExel)
        self.loadExlBrowser.setPlaceholderText("Drag & Drop!")

        loadExlBrowserbtn = QPushButton("찾아보기")
        loadExlBrowserbtn.setStyleSheet("""
            QPushButton {
                background-color: #303030;
                color: #d8d8d8;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FFA500;
            }
        """)
        loadExlBrowserbtn.clicked.connect(self.setLoadExlPath)
        loadExlLayout.addWidget(QLabel("엑셀파일"))
        loadExlLayout.addWidget(self.loadExlBrowser)
        loadExlLayout.addWidget(loadExlBrowserbtn)

        titleTableLayout = QHBoxLayout()
        self.TitleTable = QTableWidget()
        self.TitleTable.setFixedHeight(100)
        self.TitleTable.setColumnCount(4)
        self.TitleTable.setHorizontalHeaderLabels(["원본 상품명", "메인키워드", "서브키워드", "결과값"])
        self.TitleTable.setStyleSheet("background-color: #303030; color: #d8d8d8; font-weight: bold")

        mixTitleBtn = QPushButton("셔\n플")
        mixTitleBtn.setFixedSize(30, 80)
        mixTitleBtn.setStyleSheet("""
            QPushButton {
                background-color: #303030;
                color: #d8d8d8;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FFA500;
            }
        """)
        mixTitleBtn.clicked.connect(self.mixTitle)
        titleTableLayout.addWidget(self.TitleTable)
        titleTableLayout.addWidget(mixTitleBtn)

        loadKeywordsLayout = QHBoxLayout()
        self.loadKeywordsBrowser = DropLineEdit()
        self.loadKeywordsBrowser.setStyleSheet("background-color: #303030; color: #d8d8d8;")
        self.loadKeywordsBrowser.setReadOnly(True)
        self.loadKeywordsBrowser.setPlaceholderText("Drag & Drop!")
        self.loadKeywordsBrowser.pathDropped.connect(self.handleKeywords)

        loadKeywordsBtn = QPushButton("찾아보기")
        loadKeywordsBtn.setStyleSheet("""
            QPushButton {
                background-color: #303030;
                color: #d8d8d8;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FFA500;
            }
        """)
        loadKeywordsBtn.clicked.connect(self.setLoadKeywordsPath)
        loadKeywordsLayout.addWidget(QLabel("견적서"))
        loadKeywordsLayout.addWidget(self.loadKeywordsBrowser)
        loadKeywordsLayout.addWidget(loadKeywordsBtn)

        self.keyTable = QTableWidget(self)
        self.keyTable.setColumnCount(2)
        self.keyTable.setHorizontalHeaderLabels(["키워드", "기본값"])
        self.keyTable.setStyleSheet("background-color: #303030; color: #d8d8d8; font: bold")

        updateKeysBtn = QPushButton("저장")
        updateKeysBtn.setStyleSheet("""
            QPushButton {
                background-color: #303030;
                color: #d8d8d8;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FFA500;
            }
        """)
        updateKeysBtn.clicked.connect(self.updateKeywordCaches)

        resizeImgBtn = QPushButton("이미지 수정하기")
        resizeImgBtn.setStyleSheet("""
            QPushButton {
                background-color: #303030;
                color: #d8d8d8;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FFA500;
            }
        """)
        resizeImgBtn.clicked.connect(self.openImgWorkspace)


        
        mainLayout.addWidget(run1688Btn)
        mainLayout.addLayout(loadSavePathLayout)
        mainLayout.addLayout(makeNewFolderLayout)
        mainLayout.addLayout(loadExlLayout)
        mainLayout.addLayout(titleTableLayout)
        mainLayout.addLayout(loadKeywordsLayout)
        mainLayout.addWidget(self.keyTable)
        mainLayout.addWidget(updateKeysBtn)
        mainLayout.addWidget(resizeImgBtn)

        menuBar = self.menuBar()
        
        setting = menuBar.addMenu("설정")
        settingAction = QAction("공급가", self)
        settingAction.triggered.connect(self.openSetting)
        setting.addAction(settingAction)

    def makeFolder(self):
        with open("Caches/saveFolderPath.txt", "r") as file:
            self.path = f"{file.read()}/{self.folderName.text()}"
        os.makedirs(self.path, exist_ok=True)

        global imgPath
        imgPath = f"{self.path}/Images"

        os.makedirs(imgPath, exist_ok=True)

    def loadsvPath(self):
        folderPath = QFileDialog.getExistingDirectory(None, "저장경로 선택", "")
        
        with open("Caches/saveFolderPath.txt", "w") as file:
            file.write(folderPath)

        self.loadSavePath.setText(folderPath)

    def openRun1688(self):
        self.new_window = run1688()
        self.new_window.show()

    def openImgWorkspace(self):
        self.new_window = imgWorkspace()
        self.new_window.show()


    def loadExel(self):
        filePath = self.loadExlBrowser.text()
        print(filePath)
        self.dataframe = LoadExcelData().loadData(filePath)

        self.titleList = LoadExcelData().getTitleList(self.dataframe)

        self.titleDict = {"셔플": ["", "", ""]}  # 수정된 부분
        self.TitleTable.setRowCount(1)
        
        for (key, value) in self.titleDict.items():
            self.TitleTable.setItem(0, 0, QTableWidgetItem(key))
            self.TitleTable.setItem(0, 1, QTableWidgetItem(value[0]))
            self.TitleTable.setItem(0, 2, QTableWidgetItem(value[1])) 
            
    def setLoadExlPath(self):
        filePath, _ = QFileDialog.getOpenFileName(self, "파일 선택", "", "엑셀 파일 (*.xlsx *xlsm)")
        if filePath:
            self.loadExlBrowser.setText(filePath)  
            self.loadExel()


    def mixTitle(self):
        self.TitleTable.setRowCount(1)
        self.mixedTitleList = []
        for row, (key, value) in enumerate(self.titleDict.items()):
            mixedTitles = []
            firstTitle = self.TitleTable.item(row, 1).text()
            subTitles = self.TitleTable.item(row, 2).text().split(" ")
            for perm in permutations(subTitles):
                mixedTitles.append(f"{firstTitle} {' '.join(perm)}")
                self.TitleTable.setItem(row, 3, QTableWidgetItem("\n".join(mixedTitles)))

        self.mixedTitleList = mixedTitles


    def getMixedTitleList(self):
        result = []
        rows = self.TitleTable.rowCount()
        for row in range(rows):
            result.append(self.TitleTable.item(row, 3).text())
        return result
    
    def tempExcel(self):
        loadexel = LoadExcelData()
        TitleList = self.mixedTitleList

        savedKeys = Caches().getKeywordsCache()
        
        filePath = self.loadKeywordsBrowser.text()
        df = pd.read_excel(filePath, header= None, sheet_name=1)

        self.wb = load_workbook(filePath)
        self.ws = self.wb.worksheets[1]
        
        # 셔플키워드가 부족하면 리스트 복제
        while len(TitleList) < len(self.titleList):  
            TitleList = TitleList * 2

        # 필요부분만 컷팅
        if len(TitleList) >= len(self.titleList):
            TitleList = TitleList[:len(self.titleList)]

        concatTitleList = []
        for idx, mixedtitle in enumerate(TitleList): 
                mixedoptionList = (loadexel.getCombinedOptionList_fixed(loadexel.loadDataV2(self.dataframe), self.titleList[idx]))
            
                
                for option in mixedoptionList:
                    cost = loadexel.getCost(self.dataframe, self.titleList[idx], option)
                    self.calcCost(cost)
                    # 열 인덱스를 정수로 변환 (엑셀 호환)
                    NameCol = df.columns[df.iloc[4] == "상품명"]
                    NameCol = df.columns.get_loc(NameCol[0]) + 1  # 1부터 시작하는 인덱스로 변환

                    KingImgCol = df.columns[df.iloc[4] == "대표이미지 파일명"]
                    KingImgCol = df.columns.get_loc(KingImgCol[0]) + 1

                    detImgCol = df.columns[df.iloc[4] == "상세이미지 파일명"]
                    detImgCol = df.columns.get_loc(detImgCol[0]) + 1

                    supplyCol = df.columns[df.iloc[4] == "공급가"]
                    supplyCol = df.columns.get_loc(supplyCol[0]) + 1

                    coupPriceCol = df.columns[df.iloc[4] == "쿠팡 판매가"]
                    coupPriceCol = df.columns.get_loc(coupPriceCol[0]) + 1

                    
                    
                    try:
                        concatedImgname = f"{loadexel.getThumbName(self.dataframe, self.titleList[idx], str(option).split('분리선')[0])[0].split(',')[0].split('_')[0]}_상세"
                    except:
                        concatedImgname = f"{self.titleList[idx]}_상세.jpg"
                       
                    
                    
                    kingImgName = LoadExcelData().getOptionImageList(self.dataframe, self.titleList[idx], str(option).split('분리선')[0])[0]
                    if kingImgName == None:
                        kingImgName = ""


                    concatTitleList.append({"상품명" : f"{mixedtitle} {str(option).replace('분리선', ' ')}", "원가" : cost, "쿠팡 판매가" : self.coupangPrice, "공급가" : self.supplyPrice, "대표이미지 파일명": kingImgName, "상세이미지 파일명" : concatedImgname})
                    
                    size = len(concatTitleList) - 1
                    row = size + 9
                    
                    self.ws.cell(row=row, column=NameCol, value=concatTitleList[size]["상품명"])
                    self.ws.cell(row=row, column=KingImgCol, value=concatTitleList[size]["대표이미지 파일명"])
                    self.ws.cell(row=row, column=detImgCol, value=concatTitleList[size]["상세이미지 파일명"])
                    self.ws.cell(row=row, column=supplyCol, value=concatTitleList[size]["공급가"])
                    self.ws.cell(row=row, column=coupPriceCol, value=concatTitleList[size]["쿠팡 판매가"])


                    
                    for key in self.newDict.keys():
                        if key in savedKeys.keys():
                            keyCol = df.columns[df.iloc[4] == key]
                            keyCol = df.columns.get_loc(keyCol[0]) + 1
                            self.ws.cell(row=row, column=keyCol, value=savedKeys[key])           

        if self.path:
            self.wb.save(f"{self.path}/새로운 견적서.xlsx")

        else:
            self.wb.save("새로운 견적서.xlsx")    

    def calcCost(self, cost):
        data = Caches().getSupplyCache()
        exchangeRating = float(Caches().getExcRatingCache())
        coupangRating = float(Caches().getCoupPriceCache())

        cost = float(cost)

        keys = list(data.keys())
        if cost < float(keys[0]): # 3이하
            self.supplyPrice = cost * float(data[keys[0]]) * exchangeRating 
        elif cost >= float(keys[0]) and cost < float(keys[1]): # 3 ~ 5
            self.supplyPrice = cost * float(data[keys[1]]) * exchangeRating
        elif cost >= float(keys[1]) and cost < float(keys[2]): # 5 ~ 15
            self.supplyPrice = cost * float(data[keys[2]]) * exchangeRating
        elif cost >= float(keys[2]) and cost < float(keys[3]): # 15 ~ 30
            self.supplyPrice = cost * float(data[keys[3]]) * exchangeRating
        elif cost >= float(keys[3]):
            self.supplyPrice = cost * float(data[keys[4]]) * exchangeRating

        if self.supplyPrice < 100:
            self.supplyPrice = round(self.supplyPrice, -1)
        else:
            self.supplyPrice = round(self.supplyPrice, -2)

        self.coupangPrice = self.supplyPrice * coupangRating
        if self.coupangPrice < 100:
            self.coupangPrice = round(self.coupangPrice, -1)
        else:
            self.coupangPrice = round(self.coupangPrice, -2)

    def setLoadKeywordsPath(self):
        filePath, _ = QFileDialog.getOpenFileName(self, "파일 선택", "", "엑셀 파일 (*.xlsx *xlsm)")
        if filePath:
            self.loadKeywordsBrowser.setText(filePath)
            self.handleKeywords()

    def handleKeywords(self):
        filePath = self.loadKeywordsBrowser.text()
        self.loadOldKeywords()
        self.getNewKeywords()

    def loadOldKeywords(self):
        
        self.oldDict = Caches().getKeywordsCache()

        self.keyTable.setRowCount(len(self.oldDict))
        for row, (key, value) in enumerate(self.oldDict.items()):
            self.keyTable.setItem(row, 0, QTableWidgetItem(str(key)))  # 키
            self.keyTable.setItem(row, 1, QTableWidgetItem(str(value)))  # 값

    def getNewKeywords(self):
        filePath = self.loadKeywordsBrowser.text()
        df = pd.read_excel(filePath, header= None, sheet_name=1)

        required_col = df.iloc[5] == "필수"
        required_index = required_col[required_col].index.tolist()

        values = df.iloc[4, required_index].tolist()
        values.remove("대표이미지 파일명")
        values.remove("상세이미지 파일명")
        values.remove("상품명")
        values.remove("공급가")
        values.remove("쿠팡 판매가")
        values.remove("카테고리")
        values.remove("고시명")
        
        self.newDict = dict.fromkeys(values, "")

        # 키테이블 수정
        for key in self.newDict:
            if key not in self.oldDict:
                self.oldDict[key] = ""

        self.keyTable.setRowCount(len(self.oldDict))
        for row, (key, value) in enumerate(self.oldDict.items()):
            self.keyTable.setItem(row, 0, QTableWidgetItem(str(key)))  # 키
            self.keyTable.setItem(row, 1, QTableWidgetItem(str(value)))  # 값

    def updateKeywordCaches(self):
        rows = self.keyTable.rowCount()
        newDict = {}
        for row in range(rows):
            value = (self.keyTable.item(row, 0).text())
            key = (self.keyTable.item(row, 1).text())
            newDict[value] = key

        with open("Caches/keywords.txt", "w", encoding="utf-8") as file:
            json.dump(newDict, file, ensure_ascii=False, indent=4)

        self.tempExcel()

    def openSetting(self):
        self.settings = settingWindow()
        self.settings.show()   
class imgWorkspace(QWidget):
    def __init__(self):
        super().__init__()
        self.gui()

    def gui(self):
        self.setWindowTitle("이미지 합치기")
        self.setGeometry(100,100,500,500)
        self.setStyleSheet("background-color: #212121; color: #d8d8d8;")

        mainLayout = QVBoxLayout()

        loadresizeImgLayout = QHBoxLayout()
        loadresizeImgLayout.addWidget(QLabel("이미지"))

        self.resizeImgPath = DropLineEdit()
        self.resizeImgPath.setStyleSheet("background: #303030; color: #d8d8d8")
        self.resizeImgPath.setPlaceholderText("Drag & Drop!")
        self.resizeImgPath.pathDropped.connect(self.handlePath)

        loadResizeImgPath = QPushButton("찾아보기")
        loadResizeImgPath.setStyleSheet("""
            QPushButton {
                background-color: #303030;
                color: #d8d8d8;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FFA500;
            }
        """)
        loadResizeImgPath.clicked.connect(self.setLoadResizeImgPath)

        loadresizeImgLayout.addWidget(self.resizeImgPath)
        loadresizeImgLayout.addWidget(loadResizeImgPath)
        mainLayout.addLayout(loadresizeImgLayout)

        self.setMaxThumb = QLineEdit()
        self.setMaxThumb.setStyleSheet("background-color: #303030; color: #d8d8d8;")
        self.setMaxThumb.setPlaceholderText("상세페이지 최대 이미지 수")

        self.scroll_area = QScrollArea()
        self.scroll_area.setMinimumHeight(160)
        self.scroll_widget = QWidget()
        self.ThumbLayout = QHBoxLayout(self.scroll_widget)


        self.scroll_widget.setLayout(self.ThumbLayout)
        self.scroll_area.setWidget(self.scroll_widget)
        self.scroll_area.setWidgetResizable(True)

        mainLayout.addWidget(self.setMaxThumb)
        mainLayout.addWidget(self.scroll_area)

        self.concatBtn = QPushButton("조합")
        self.concatBtn.setStyleSheet("""
            QPushButton {
                background-color: #303030;
                color: white;
            }
            QPushButton:hover {
                background-color: #FFA500;
            }
        """)
        self.concatBtn.clicked.connect(self.concatImg)

        mainLayout.addWidget(self.concatBtn)
        self.setLayout(mainLayout)


    def concatImg(self):
        topImg = Image.open("상단이미지/상단이미지.001.png")
        self.savePath = imgPath


        for file_no in self.thumbImgsDict.keys():
            selected_thumbnails = [self.checkboxList[file_no][cb] for cb in self.checkboxList[file_no] if cb.isChecked()]
            option_images = self.optionImgsDict.get(file_no, [])
            if len(option_images) == 0: #옵션이 비어있으면
                continue
        
            # 옵션 이미지 불러오기
            deimages = [Image.open(img) for img in option_images]
            
            img_width, img_height = deimages[0].size
            
            # 상단 이미지 사이즈 조절 (가로 크기를 기준으로 조정)
            new_width = img_width * 2 + 15
            width_percent = new_width / float(topImg.size[0])
            new_height = int(float(topImg.size[1]) * width_percent)
            topImg = topImg.resize((new_width, new_height), Image.LANCZOS)
            
            # 체크된 썸네일 이미지 가져오기
            thimages = [Image.open(img) for img in selected_thumbnails]
            
            # 썸네일 이미지 크기 조정 (topImg의 너비에 맞춤)
            thumb_width = topImg.width
            for i in range(len(thimages)):
                scale = thumb_width / thimages[i].width
                new_thumb_height = int(thimages[i].height * scale)
                thimages[i] = thimages[i].resize((thumb_width, new_thumb_height), Image.LANCZOS)
            
            # 썸네일 높이 계산
            total_thumb_height = sum(img.height for img in thimages)
            
            # 옵션 이미지 배열 설정 (2열 배치)
            cols = 2
            rows = math.ceil(len(deimages) / cols)
            gap = 15
            text_height = 15  # 텍스트 높이
            
            de_width = (img_width + gap) * cols - gap
            de_height = (img_height + text_height + gap) * rows - gap
            
            # 최종 캔버스 크기 계산
            final_width = max(topImg.width, de_width)
            final_height = topImg.height + total_thumb_height + de_height
            final_img = Image.new("RGB", (final_width, final_height), "white")
            
            # 상단 이미지 배치
            y_offset = 0
            final_img.paste(topImg, (0, y_offset))
            y_offset += topImg.height
            
            # 썸네일 배치
            for img in thimages:
                final_img.paste(img, (0, y_offset))
                y_offset += img.height
            
            # 옵션 이미지 배치 (2열)
            draw = ImageDraw.Draw(final_img)
            try:
                font = ImageFont.load_default()
            except:
                print("no font")
            
            for idx, img in enumerate(deimages):
                x_offset = (idx % cols) * (img_width + gap)
                y_position = y_offset + (idx // cols) * (img_height + text_height + gap)
                
                final_img.paste(img, (x_offset, y_position))
                
                # 파일명 텍스트 추가 (예제용으로 빈 문자열)
                filename = ""
                text_x = x_offset + 10
                text_y = y_position + img_height + 5
                draw.text((text_x, text_y), filename, font=font, fill="black")
            
            # 이미지 크기 조정 후 저장
            final_img.save(f"{self.savePath}/{file_no}_상세.jpg")

        self.getdetailImgs()
        self.clear_layout(self.ThumbLayout)

    def clear_layout(self, layout):
        # 레이아웃의 모든 위젯 제거
        if layout is not None:
            while layout.count():
                item = layout.takeAt(0)  # 첫 번째 아이템을 가져옴
                if item.widget():
                    item.widget().deleteLater()
    
    def reset_scroll_widget(self):
        # scroll_widget을 새로 생성하여 레이아웃을 초기화
        self.scroll_widget = QWidget()  # 새로운 scroll_widget 생성
        self.ThumbLayout = QHBoxLayout(self.scroll_widget)  # 새로운 레이아웃을 설정
        
        # scroll_widget을 QScrollArea에 다시 설정
        self.scroll_area.setWidget(self.scroll_widget)
    
    def handlePath(self):
        self.resizeImg()

        # 썸네일 추가할 때 상품번호 기준으로 행 정렬
        self.reset_scroll_widget()
        grid_layout = QGridLayout()
        
        row = 0
        self.checkboxList = {}  # 상품번호별 체크박스 저장
        
        for no in sorted(self.thumbImgsDict.keys(), key=int):
            col = 0
            self.checkboxList[no] = {}  # 상품번호별 딕셔너리 생성
            
            for img in self.thumbImgsDict[no]:
                checkbox = QCheckBox()
                pixmap = QPixmap(img)
                icon = QIcon(pixmap)
                checkbox.setIcon(icon)
                checkbox.setIconSize(QSize(120,120))
                checkbox.setChecked(True)
                self.checkboxList[no][checkbox] = img  # 상품번호별 체크박스 저장
                grid_layout.addWidget(checkbox, row, col)
                col += 1
            row += 1
        
        self.ThumbLayout.addLayout(grid_layout)

    def setLoadResizeImgPath(self):
        folderPath = QFileDialog.getExistingDirectory(None, "이미지 폴더 선택", "")
        self.resizeImgPath.setText(folderPath)
        self.handlePath()

    def getdetailImgs(self):
        '''옵션이미지를 다운폴더에 복제'''
        folderPath = self.resizeImgPath.text()

        imgFolder = os.path.join(folderPath, "옵션")
        # 이미지 확장자 정의
        image_exts = ['.png', '.jpg', '.jpeg', '.bmp', '.gif']

        # 경로 리스트 만들기
        images = [
            os.path.join(imgFolder, file)
            for file in os.listdir(imgFolder)
            if os.path.splitext(file)[1].lower() in image_exts
        ]

        for img in images:
            fileName = os.path.basename(img)
            filePath = os.path.join(imgPath, fileName)
            img = Image.open(img)
            img.save(filePath)
        

    def resizeImg(self):
        folderPath = self.resizeImgPath.text()
        
        self.thumbImgsDict = {}
        self.optionImgsDict = {}

        for folder, img_dict, size in [("썸네일", self.thumbImgsDict, (780, 500)), ("옵션", self.optionImgsDict, (500, 500))]:
            img_folder = os.path.join(folderPath, folder)
            if not os.path.exists(img_folder):
                continue
            
            image_files = [
                os.path.join(img_folder, f) for f in os.listdir(img_folder)
                if os.path.isfile(os.path.join(img_folder, f)) and os.path.splitext(f)[1].lower() in {".jpg", ".jpeg", ".png", ".bmp", ".gif", ".webp"}
            ]
            
            for image in image_files:
                file_name = os.path.basename(image)
                file_no = file_name.split("_")[0]
                if file_no not in img_dict:
                    img_dict[file_no] = []
                img_dict[file_no].append(self.resizeImgTool(image, *size))


    def resizeImgTool(self, img_path, min_width, min_height):
        
        img = Image.open(img_path)
        original_width, original_height = img.size
        
        # 기존 크기가 최소 크기보다 크면 변경할 필요 없음
        if original_width >= min_width and original_height >= min_height:
            return img_path

        # 리사이징 비율 계산
        scale_w = min_width / original_width
        scale_h = min_height / original_height
        scale = max(scale_w, scale_h)
        
        # 새로운 크기 계산
        new_width = int(original_width * scale)
        new_height = int(original_height * scale)

        # 이미지 리사이징 후 원본 파일 덮어쓰기
        resized_img = img.resize((new_width, new_height), Image.LANCZOS)
        resized_img.save(img_path)  # 원본 파일에 덮어쓰기
        return img_path

class settingWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.settingGUI()

    def settingGUI(self):
        self.setWindowTitle("공급가")
        self.setGeometry(100,100,300,430)
        self.setStyleSheet("background-color: #212121; color: #d8d8d8;")

        mainLayout = QVBoxLayout()
           
        exchangeLayout = QHBoxLayout()
        self.exchangeRating = QLineEdit()
        self.exchangeRating.setStyleSheet("background-color: #303030; color: #d8d8d8;")
        exchangeLayout.addWidget(QLabel("환율"))
        exchangeLayout.addWidget(self.exchangeRating)

        coupangpriceLayout = QHBoxLayout()
        self.coupangPrice = QLineEdit()
        self.coupangPrice.setStyleSheet("background-color: #303030; color: #d8d8d8;")
        coupangpriceLayout.addWidget(QLabel("쿠팡 판매가"))
        coupangpriceLayout.addWidget(self.coupangPrice)

        coupangSupplyPriceLayout = QVBoxLayout()
        self.coupangSupplyTable = QTableWidget()
        self.coupangSupplyTable.setColumnCount(2)
        self.coupangSupplyTable.setHorizontalHeaderLabels(["기준", "배수"])
        
        self.coupangSupplyTable.setStyleSheet("background-color: #303030; color: #d8d8d8;")
        coupangSupplyPriceLayout.addWidget(QLabel("쿠팡 공급가"))
        coupangSupplyPriceLayout.addWidget(self.coupangSupplyTable)

        saveBtn = QPushButton("저장")
        saveBtn.setStyleSheet("""
            QPushButton {
                background-color: #303030;
                color: #d8d8d8;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FFA500;
            }
        """)
        saveBtn.clicked.connect(self.saveSupplyCaches)

        mainLayout.addLayout(exchangeLayout)
        mainLayout.addLayout(coupangpriceLayout)
        mainLayout.addLayout(coupangSupplyPriceLayout)
        mainLayout.addWidget(saveBtn)
        mainLayout.addWidget(QLabel("판매가 = 공급가 x 판매배수"))
        mainLayout.addWidget(QLabel("공급가 = 원가 x 공급배수 x 환율"))

        self.loadSupplyCaches()
        self.setLayout(mainLayout)

    def loadSupplyCaches(self):
        data = Caches().getSupplyCache()

        self.coupangSupplyTable.setRowCount(len(data))
        for row, (key, value) in enumerate(data.items()):
            self.coupangSupplyTable.setItem(row, 0, QTableWidgetItem(key))
            self.coupangSupplyTable.setItem(row, 1, QTableWidgetItem(value))

        data = Caches().getCoupPriceCache()
        self.coupangPrice.setText(data)

        data = Caches().getExcRatingCache()
        self.exchangeRating.setText(data)

    def saveSupplyCaches(self):
        rows = self.coupangSupplyTable.rowCount()
        newSupData = {}
        for row in range(rows):
            newKey = self.coupangSupplyTable.item(row, 0)
            newValue = self.coupangSupplyTable.item(row, 1)
            newSupData[newKey.text()] = newValue.text()

        newCoupPriceData = self.coupangPrice.text()
        newExcRatingData = self.exchangeRating.text()
        newData = [newSupData, newExcRatingData, newCoupPriceData]

        with open("Caches/supply.txt", "w", encoding="utf-8") as file:
            json.dump(newData, file, ensure_ascii=False, indent=4)
class loginWindow(QWidget):

    def __init__(self):
        super().__init__()
        self.loginUI()

    def loginUI(self):
        self.setWindowTitle("로그인")
        self.setGeometry(300,300,300,150)
        self.setStyleSheet("background-color: black; color: white")

        mainLayout = QVBoxLayout()

        self.permission = QLabel()
        self.macLabel = QLabel()
        self.pwLine = QLineEdit()
        self.loginBtn = QPushButton("로그인")
        self.loginBtn.clicked.connect(self.login)

        self.adminPw = QPushButton("엑셀암호찾기")
        self.adminPw.clicked.connect(self.admin)
        self.adminPw.setEnabled(False)

        mainLayout.addWidget(self.permission)
        mainLayout.addWidget(self.macLabel)
        mainLayout.addWidget(self.pwLine)
        mainLayout.addWidget(self.loginBtn)
        mainLayout.addWidget(self.adminPw)

        self.setLayout(mainLayout)
        

    def getMac(self):
        # 암호화된 엑셀 파일 열기
        file_path = "Caches/login.xlsx"
        password = "asasas2@"  # 엑셀 파일의 비밀번호 입력 asasas2@ fhkswh11

        myMac = get_mac_address()

        try:
            with open(file_path, "rb") as file:
                decrypted = BytesIO()
                office_file = msoffcrypto.OfficeFile(file)
                office_file.load_key(password)  # 비밀번호 입력
                office_file.decrypt(decrypted)  # 복호화된 내용을 메모리에 저장

            # 복호화된 데이터로 pandas에서 읽기
            decrypted.seek(0)
            df = pd.read_excel(decrypted, engine="openpyxl")
        except:
            self.errorMsg()
            
        dflist = df.values.tolist()
        
        developKey = "18:3e:ef:c3:a4:b5"
        
        try:
            key = dflist[0][1]
            
            if str(key) != developKey:
                raise Exception()
        except:
            self.errorMsg()
            self.close()


        for user in dflist:
            if user[1] == myMac:
                myId = user
                break
            else:
                myId = None

        if myId:
            self.permission.setText(f"내 권한 : {myId[0]}")
            self.macLabel.setText(f"내 아이디 : {myId[1]}")
            self.myPw = myId[2]
            if myId[0] == "관리자" or myId[0] == "개발자":
                self.adminPw.setDisabled(False)
        elif myId == None:
            self.permission.setText(f"내 권한 : 미등록유저")
            self.macLabel.setText(f"내 아이디 : {myMac}")
            self.loginBtn.setEnabled(False)
        else:
            return False

    def admin(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)  # 아이콘 종류: Information, Warning, Critical, Question
        msg.setWindowTitle("열기암호찾기")
        msg.setText(f"열기암호 : asasas2@") #admin3230192
        msg.setStandardButtons(QMessageBox.Ok)  # 버튼 추가
        msg.exec()  # 메시지 박스 실행

    
    def errorMsg(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)  # 아이콘 종류: Information, Warning, Critical, Question
        msg.setWindowTitle("에러메세지")
        msg.setText("유저 데이터가 유실되었습니다. 관리자에게 문의하세요.")
        msg.setStandardButtons(QMessageBox.Ok)  # 버튼 추가
        msg.exec()  # 메시지 박스 실행
        

    def login(self):
        Pw = self.pwLine.text()
        if Pw == str(self.myPw):
            self.window = MainWindow()
            self.window.show()
            self.close()
            
        else:
            return False

if __name__ == "__main__":
    if getattr(sys, 'frozen', False):  # PyInstaller 실행 파일 여부 확인
        exe_dir = os.path.dirname(sys.executable)
        os.chdir(exe_dir)  # 실행 파일이 있는 곳으로 작업 디렉토리 변경
    
    app = QApplication([])
    '''
    login_Window = loginWindow()
    login_Window.show()
    login_Window.getMac()
    '''
    window = MainWindow()
    window.show()
    
    app.exec()